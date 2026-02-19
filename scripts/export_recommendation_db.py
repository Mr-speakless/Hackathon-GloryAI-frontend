import json
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / "AI_Skin_Product_Recommendation_Database.xlsx"
OUTPUT_PATH = ROOT / "GloryAI_frontend" / "src" / "data" / "productRecommendationDb.json"


def normalize_score_rule(value):
    text = str(value or "").strip()
    if text.startswith(">"):
        return ">80"
    if text.startswith("<"):
        return "<80"
    return ""


def normalize_level(value):
    text = str(value or "").strip()
    if text.lower() == "basic":
        return "Basic"
    if text.lower() == "advanced":
        return "Advanced"
    return text


def normalize_price(value):
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return int(value) if float(value).is_integer() else float(value)
    text = str(value).strip()
    if not text:
        return ""
    try:
        number = float(text)
        return int(number) if number.is_integer() else number
    except ValueError:
        return text


def get_image_file(target_concern, level):
    suffix = "1" if level == "Basic" else "2" if level == "Advanced" else ""
    concern = str(target_concern or "").strip()
    return f"{concern}-{suffix}.png" if concern and suffix else ""


def normalize_row(row):
    # New sheet columns: A(image), B(brand), C(product), D(price), E(concern), F(level), G(score rule), H(url)
    brand = str(row[1] or "").strip()
    name = str(row[2] or "").strip()
    price = normalize_price(row[3])
    target_concern = str(row[4] or "").strip()
    level = normalize_level(row[5])
    score_rule = normalize_score_rule(row[6])
    product_url = str(row[7] or "").strip()

    return {
        "brand": brand,
        "name": name,
        "price": price,
        "targetConcern": target_concern,
        "level": level,
        "scoreRule": score_rule,
        "imageFile": get_image_file(target_concern, level),
        "productUrl": product_url,
    }


def export_db():
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"XLSX not found: {XLSX_PATH}")

    workbook = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    products = []
    for row_idx in range(2, sheet.max_row + 1):
        row = [sheet.cell(row=row_idx, column=col).value for col in range(1, 9)]
        product = normalize_row(row)
        if product["name"]:
            products.append(product)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(products, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Exported {len(products)} products -> {OUTPUT_PATH}")


if __name__ == "__main__":
    export_db()
